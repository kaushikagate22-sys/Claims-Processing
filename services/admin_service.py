"""
services.admin_service
====================
Powers the in-app Admin screen: upload/edit the policy document (auto-recompile
rules), upload the structured-data CSVs (auto-reseed the database), and read back
the current policy + records.

Crucially it **invalidates the relevant caches** after a change, so the live
pipeline picks up new policy values and records without a server restart.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List

import yaml

from config.settings import get_settings, load_rules
from core.tools import record_store
from db.database import session_scope
from db.repository import list_history, list_policies, replace_history, replace_policies
from domain.policy_compiler import PolicyCompiler

POLICY_REQUIRED = ["policy_number"]
HISTORY_REQUIRED = ["policy_number"]


class AdminService:
    def __init__(self) -> None:
        self.settings = get_settings()

    # --- policy document --------------------------------------------------
    def get_policy(self) -> Dict[str, Any]:
        text = self.settings.master_policy.read_text(encoding="utf-8")
        rules = load_rules()
        return {
            "text": text,
            "claim_types": rules.get("claim_types", {}),
            "rules_count": len(rules.get("rules", [])),
        }

    def save_policy(self, text: str) -> Dict[str, Any]:
        self.settings.master_policy.write_text(text, encoding="utf-8")
        result = PolicyCompiler().compile()  # writes rules.yaml (+ backup)
        load_rules.cache_clear()              # so live decisions use new rules
        return {
            "diff": result["diff"],
            "claim_types": result["compiled"]["claim_types"],
            "mode": result["mode"],
        }

    # --- structured records ----------------------------------------------
    def save_policies_csv(self, raw: bytes, filename: str = "") -> Dict[str, Any]:
        raw = self._to_csv_bytes(raw, filename)
        rows = self._validate_csv(raw, POLICY_REQUIRED, "policies")
        self.settings.policies_db.write_bytes(raw)
        record_store.clear_cache()            # so validation reads new records
        with session_scope() as s:
            count = replace_policies(s, str(self.settings.policies_db))
        return {"count": count, "rows": len(rows)}

    def save_history_csv(self, raw: bytes, filename: str = "") -> Dict[str, Any]:
        raw = self._to_csv_bytes(raw, filename)
        rows = self._validate_csv(raw, HISTORY_REQUIRED, "claims history")
        self.settings.claims_history_db.write_bytes(raw)
        record_store.clear_cache()
        with session_scope() as s:
            count = replace_history(s, str(self.settings.claims_history_db))
        return {"count": count, "rows": len(rows)}

    @staticmethod
    def _to_csv_bytes(raw: bytes, filename: str = "") -> bytes:
        """Accept Excel (.xlsx) as well as CSV — convert xlsx to CSV bytes."""
        is_xlsx = filename.lower().endswith((".xlsx", ".xlsm")) or raw[:2] == b"PK"
        if not is_xlsx:
            return raw
        from openpyxl import load_workbook  # lazy import
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in row])
        return buf.getvalue().encode("utf-8")

    def get_records(self) -> Dict[str, Any]:
        with session_scope() as s:
            return {"policies": list_policies(s), "history": list_history(s)}

    # --- master / system-of-record files (editable inputs) ----------------
    def _master_paths(self) -> Dict[str, Any]:
        s = self.settings
        return {
            "employee": s.employee_master, "dealer": s.dealer_master,
            "machine": s.machine_master, "parts": s.parts_master, "transit": s.transit_master,
        }

    MASTER_LABELS = {
        "employee": "Employee master", "dealer": "Dealer / retail master",
        "machine": "Machine / serial master", "parts": "Parts master", "transit": "Transit / consignment master",
    }

    def get_masters(self) -> Dict[str, Any]:
        out = []
        for kind, path in self._master_paths().items():
            columns: List[str] = []
            rows: List[Dict[str, str]] = []
            if path.exists():
                with open(path, newline="", encoding="utf-8") as fh:
                    reader = csv.DictReader(fh)
                    columns = list(reader.fieldnames or [])
                    rows = [dict(r) for r in reader]
            out.append({"kind": kind, "label": self.MASTER_LABELS.get(kind, kind),
                        "columns": columns, "rows": rows, "count": len(rows)})
        return {"masters": out}

    def save_master(self, kind: str, raw: bytes, filename: str = "") -> Dict[str, Any]:
        paths = self._master_paths()
        if kind not in paths:
            raise ValueError(f"Unknown master '{kind}'. Expected one of {list(paths)}.")
        raw = self._to_csv_bytes(raw, filename)
        # validate it parses and has a header + at least one column
        rows = self._validate_csv(raw, [], self.MASTER_LABELS.get(kind, kind))
        path = paths[kind]
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(raw)  # master_data tool reloads automatically (mtime change)
        return {"kind": kind, "rows": len(rows)}

    # --- claim types (config-driven) -------------------------------------
    OPERATORS = ["exists", "not_exists", "truthy", "falsy", "eq", "ne",
                 "gt", "gte", "lt", "lte", "in", "not_in", "contains", "regex"]
    SEVERITIES = ["reject", "hold", "escalate", "info"]
    # derived facts the engine computes (selectable in rule editor besides fields)
    DERIVED_FIELDS = ["claim_amount", "excluded", "is_duplicate", "high_value",
                      "needs_escalation", "complete", "handover_open",
                      "dispatch_before_inspection", "manager_approval_ok",
                      # validation / master-data flags
                      "dealer_known", "serial_known", "warranty_active", "part_known",
                      "consignment_known", "employee_known", "employee_active", "within_reimb_limit"]

    def get_types(self) -> Dict[str, Any]:
        from domain.type_config import extraction_schema_for
        rules = load_rules()
        out = {}
        for ct, cfg in rules.get("claim_types", {}).items():
            schema = extraction_schema_for(ct)
            field_options = sorted(set(list(schema.keys()) + self.DERIVED_FIELDS))
            out[ct] = {
                "label": cfg.get("label", ct),
                "required_fields": cfg.get("required_fields", []),
                "exclusion_keywords": cfg.get("exclusion_keywords", []),
                "escalate_keywords": cfg.get("escalate_keywords", []),
                "duplicate_key": cfg.get("duplicate_key"),
                "high_value_threshold": cfg.get("high_value_threshold"),
                "settlement_cap": cfg.get("settlement_cap"),
                "checks": cfg.get("checks", []),
                "visual_checks": cfg.get("visual_checks", []),
                "extraction_schema": schema,
                "field_options": field_options,
            }
        return {
            "types": out,
            "operators": self.OPERATORS,
            "severities": self.SEVERITIES,
        }

    def save_type_fields(self, ctype: str, fields: Dict[str, str]) -> Dict[str, Any]:
        rpath = self.settings.rules_path
        data = yaml.safe_load(rpath.read_text(encoding="utf-8")) or {}
        cts = data.setdefault("claim_types", {})
        if ctype not in cts:
            raise ValueError(
                f"Claim type '{ctype}' is not configured yet."
            )
        clean = {str(k).strip(): str(v).strip()
                 for k, v in (fields or {}).items() if str(k).strip()}
        cts[ctype]["extraction_schema"] = clean
        rpath.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")
        load_rules.cache_clear()
        return {"type": ctype, "fields": clean}

    def save_type_visual(self, ctype: str, criteria: List[str]) -> Dict[str, Any]:
        """Persist a type's visual-validation criteria (the photo checks the visual
        agent reports against) to rules.yaml."""
        rpath = self.settings.rules_path
        data = yaml.safe_load(rpath.read_text(encoding="utf-8")) or {}
        cts = data.setdefault("claim_types", {})
        if ctype not in cts:
            raise ValueError(f"Claim type '{ctype}' is not configured.")
        clean = [str(c).strip() for c in (criteria or []) if str(c).strip()][:6]
        cts[ctype]["visual_checks"] = clean
        rpath.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")
        load_rules.cache_clear()
        return {"type": ctype, "visual_checks": clean}

    def save_type_rules(self, ctype: str, checks: List[Dict[str, Any]],
                        config: Dict[str, Any] | None = None) -> Dict[str, Any]:
        """Persist a type's validation checks (+ optional config knobs) to
        rules.yaml. This is the source of truth for live decisions."""
        rpath = self.settings.rules_path
        data = yaml.safe_load(rpath.read_text(encoding="utf-8")) or {}
        cts = data.setdefault("claim_types", {})
        if ctype not in cts:
            raise ValueError(f"Claim type '{ctype}' is not configured.")

        clean_checks = []
        seen_ids = set()
        for i, c in enumerate(checks or []):
            cid = str(c.get("id") or "").strip()
            field = str(c.get("field") or "").strip()
            op = str(c.get("op") or "").strip()
            sev = str(c.get("severity") or "").strip()
            if not (cid and field and op):
                continue
            if op not in self.OPERATORS:
                raise ValueError(f"Rule '{cid}': unknown operator '{op}'.")
            if sev not in self.SEVERITIES:
                raise ValueError(f"Rule '{cid}': severity must be one of {self.SEVERITIES}.")
            if cid in seen_ids:
                raise ValueError(f"Duplicate rule id '{cid}'.")
            seen_ids.add(cid)
            rule: Dict[str, Any] = {"id": cid, "field": field, "op": op, "severity": sev,
                                    "category": str(c.get("category") or "eligibility").strip(),
                                    "message": str(c.get("message") or "").strip()}
            val = c.get("value")
            if val not in (None, ""):
                # numeric ops keep numbers; others keep as given
                if op in ("gt", "gte", "lt", "lte", "eq", "ne"):
                    try:
                        val = float(val)
                    except (TypeError, ValueError):
                        pass
                rule["value"] = val
            clean_checks.append(rule)

        cts[ctype]["checks"] = clean_checks

        if config:
            for key in ("required_fields", "exclusion_keywords", "escalate_keywords"):
                if key in config and isinstance(config[key], list):
                    cts[ctype][key] = [str(x).strip() for x in config[key] if str(x).strip()]
            if "duplicate_key" in config:
                cts[ctype]["duplicate_key"] = (config["duplicate_key"] or None)
            for key in ("high_value_threshold", "settlement_cap"):
                if key in config:
                    v = config[key]
                    try:
                        cts[ctype][key] = float(v) if v not in (None, "") else None
                    except (TypeError, ValueError):
                        cts[ctype][key] = None

        rpath.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")
        load_rules.cache_clear()
        return {"type": ctype, "checks": clean_checks}

    # --- helpers ----------------------------------------------------------
    @staticmethod
    def _validate_csv(raw: bytes, required: List[str], label: str) -> List[Dict[str, str]]:
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError as e:
            raise ValueError(f"The {label} file must be UTF-8 CSV text.") from e
        reader = csv.DictReader(io.StringIO(text))
        cols = reader.fieldnames or []
        missing = [c for c in required if c not in cols]
        if missing:
            raise ValueError(
                f"The {label} CSV is missing required column(s): {', '.join(missing)}."
            )
        rows = list(reader)
        if not rows:
            raise ValueError(f"The {label} CSV has no data rows.")
        return rows
